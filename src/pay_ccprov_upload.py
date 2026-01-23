import csv
from openpyxl import load_workbook
from psycopg2.extras import execute_values
# from db.easebase_conn import easebase_conn

def _noneify(v):
    # Convert blanks to None; keep numbers/dates as-is
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v

def get_db_connection():
    try:
        db_credentials = json.loads(os.getenv("DB_CREDENTIALS"))  # Retrieve and parse DB credentials
        conn = psycopg2.connect(
            dbname=db_credentials["database"],
            user=db_credentials["user"],
            password=db_credentials["password"],
            host=db_credentials["host"],
            port=db_credentials["port"]
        )
        return conn
    except Exception as e:
        print(f"🚨 Database connection error: {e}")
        raise

def _read_csv_rows(path):
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        header = next(r, None)
        if header is None:
            return [], []
        rows = []
        for row in r:
            rows.append([_noneify(v) for v in row])
        return header, rows

def _read_xlsx_rows(path, header_row=1):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    # Read header
    header = [c.value for c in ws[header_row]]
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # skip empty rows
        if row is None:
            continue
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        rows.append([_noneify(v) for v in row])
    return header, rows

def upload_to_postgres(ccprov_csv_path, ccstaff_csv_path, labor_xlsx_path):
    conn = get_db_connection()
    cursor = conn.cursor()

    table_name = 'app.clinic_ccprov'
    lab_table_name = 'app.clinic_labor_costs'
    staff_table_name = 'app.clinic_ccstaff'

    cursor.execute('truncate app.clinic_ccprov')
    cursor.execute('truncate app.clinic_ccstaff')
    cursor.execute('truncate app.clinic_labor_costs')
    conn.commit()

    batch_size = 100

    # ccprov
    _, ccprov_rows = _read_csv_rows(ccprov_csv_path)
    insert_query = f"""INSERT INTO {table_name} (
        ee_id, employee_name, shift_date, day, pay_type, reg_hours, ot1_hours, ot2_hours,
        unpaid_hours, time_in, time_out, cc1, cc2, reg_charge_rate, ot_charge_rate,
        reg_charge_amount, ot_charge_amount, total_charge_amount, reg_pay_rate, ot_pay_rate,
        reg_paid, ot_paid, total_pay_amount
    ) VALUES %s ON CONFLICT DO NOTHING;"""

    buf = []
    for row in ccprov_rows:
        buf.append(tuple(row))
        if len(buf) == batch_size:
            execute_values(cursor, insert_query, buf)
            conn.commit()
            buf = []
    if buf:
        execute_values(cursor, insert_query, buf)
        conn.commit()

    # ccstaff
    _, ccstaff_rows = _read_csv_rows(ccstaff_csv_path)
    staff_insert_query = f"""INSERT INTO {staff_table_name} (
        ee_id, employee_name, shift_date, day, pay_type, reg_hours, ot1_hours, ot2_hours,
        unpaid_hours, time_in, time_out, cc1, cc2, reg_charge_rate, ot_charge_rate,
        reg_charge_amount, ot_charge_amount, total_charge_amount, reg_pay_rate, ot_pay_rate,
        reg_paid, ot_paid, total_pay_amount
    ) VALUES %s ON CONFLICT DO NOTHING;"""

    buf = []
    for row in ccstaff_rows:
        buf.append(tuple(row))
        if len(buf) == batch_size:
            execute_values(cursor, staff_insert_query, buf)
            conn.commit()
            buf = []
    if buf:
        execute_values(cursor, staff_insert_query, buf)
        conn.commit()

    # labor (xlsx)
    _, labor_rows = _read_xlsx_rows(labor_xlsx_path, header_row=1)
    lab_insert_query = f"""INSERT INTO {lab_table_name} (
        company, ee_id, cntlast, cntfirst, cntdept, cndarea, last_check_dt, cc1, cc2,
        reg_hours, reg_amount, ot_hours, ot_amount, bonus_amount, other_amount,
        suta, futa, ss_tax, mcare_tax, other_tax, ret_ben, med_ben, dent_ben, vis_ben, oth_ben
    ) VALUES %s ON CONFLICT DO NOTHING;"""

    buf = []
    for row in labor_rows:
        buf.append(tuple(row))
        if len(buf) == batch_size:
            execute_values(cursor, lab_insert_query, buf)
            conn.commit()
            buf = []
    if buf:
        execute_values(cursor, lab_insert_query, buf)
        conn.commit()

    cursor.close()
    conn.close()
