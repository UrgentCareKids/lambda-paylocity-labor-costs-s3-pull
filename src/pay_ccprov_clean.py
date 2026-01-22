import os
import csv
from openpyxl import load_workbook

def clean_ccprov_and_ccstaff(ccprov_paths, ccstaff_paths, out_dir="/tmp"):
    """
    ccprov_paths: list[str] local .xlsx paths for ccprov files
    ccstaff_paths: list[str] local .xlsx paths for ccstaff files
    returns: (ccprov_csv_path, ccstaff_csv_path)
    """

    def _read_xlsx_rows(path: str, skiprows: int = 5):
        # data_only=True uses computed values for formulas (if saved)
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active  # assumes first sheet; adjust if needed

        # openpyxl rows are 1-indexed
        start_row = skiprows + 1
        rows_iter = ws.iter_rows(min_row=start_row, values_only=True)

        # Get header row (first row after skiprows)
        header = next(rows_iter, None)
        if header is None:
            return [], []

        header = list(header)
        # Drop 3rd column (index 2) if present
        if len(header) > 2:
            header.pop(2)

        cleaned_rows = []
        for r in rows_iter:
            if r is None:
                continue
            row = list(r)
            # Drop 3rd column (index 2) if present
            if len(row) > 2:
                row.pop(2)

            # Optional: skip completely empty rows
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                continue

            cleaned_rows.append(row)

        return header, cleaned_rows

    def _merge_and_write_csv(paths, out_path):
        merged_header = None
        merged_rows = []

        for p in paths:
            header, rows = _read_xlsx_rows(p, skiprows=5)
            if merged_header is None:
                merged_header = header
            else:
                # Basic safety: ensure headers match; if not, still proceed but you can enforce
                if header != merged_header:
                    raise ValueError(f"Header mismatch in file {p}")

            merged_rows.extend(rows)

        if merged_header is None:
            raise ValueError("No data found in provided Excel files")

        with open(out_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(merged_header)
            w.writerows(merged_rows)

    ccprov_out = os.path.join(out_dir, "cleaned_clinic_ccprov.csv")
    ccstaff_out = os.path.join(out_dir, "cleaned_clinic_ccstaff.csv")

    _merge_and_write_csv(ccprov_paths, ccprov_out)
    _merge_and_write_csv(ccstaff_paths, ccstaff_out)

    return ccprov_out, ccstaff_out
